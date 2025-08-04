#!/usr/bin/env python3
"""Export GitLab analytics to various formats including Excel."""

import sys
import argparse
from pathlib import Path
from datetime import datetime
import json

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.api import GitLabClient
from src.services import GitLabAnalytics
from src.services.analytics_advanced import AdvancedAnalytics
from src.utils import Config, setup_logging, get_logger
from src.utils.logger import Colors

logger = get_logger(__name__)


def export_to_excel(data: dict, output_path: str):
    """Export analytics data to Excel format.
    
    Args:
        data: Analytics data
        output_path: Output file path
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Overview sheet
        ws_overview = wb.create_sheet("Overview")
        ws_overview.append(["GitLab Analytics Report"])
        ws_overview.append(["Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_overview.append([])
        
        # Style the header
        ws_overview['A1'].font = Font(size=16, bold=True)
        ws_overview['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_overview['A1'].font = Font(color="FFFFFF", size=16, bold=True)
        
        # Project or group data
        if 'project' in data:
            # Single project
            project = data['project']
            ws_overview.append(["Project Information"])
            ws_overview.append(["Name:", project['name']])
            ws_overview.append(["Path:", project['path']])
            ws_overview.append(["Created:", project['created_at']])
            ws_overview.append(["Last Activity:", project['last_activity_at']])
            ws_overview.append([])
            
            # Commits sheet
            if 'commits' in data:
                ws_commits = wb.create_sheet("Commits")
                commits_data = data['commits']
                
                # Summary
                ws_commits.append(["Commit Statistics"])
                ws_commits.append(["Total Commits (30 days):", commits_data.get('total', 0)])
                ws_commits.append(["Average per Day:", commits_data.get('average_per_day', 0)])
                ws_commits.append([])
                
                # By author
                if commits_data.get('by_author'):
                    ws_commits.append(["Commits by Author"])
                    ws_commits.append(["Author", "Commits"])
                    for author, count in sorted(
                        commits_data['by_author'].items(), 
                        key=lambda x: x[1], 
                        reverse=True
                    ):
                        ws_commits.append([author, count])
            
            # Issues sheet
            if 'issues' in data:
                ws_issues = wb.create_sheet("Issues")
                issues_data = data['issues']
                
                ws_issues.append(["Issue Statistics"])
                ws_issues.append(["Total Issues:", issues_data.get('total', 0)])
                ws_issues.append(["Open Issues:", issues_data.get('open', 0)])
                ws_issues.append(["Closed Issues:", issues_data.get('closed', 0)])
                ws_issues.append(["Closure Rate:", f"{issues_data.get('closure_rate', 0)*100:.1f}%"])
                ws_issues.append([])
                
                # By label
                if issues_data.get('by_label'):
                    ws_issues.append(["Issues by Label"])
                    ws_issues.append(["Label", "Count"])
                    for label, count in sorted(
                        issues_data['by_label'].items(), 
                        key=lambda x: x[1], 
                        reverse=True
                    ):
                        ws_issues.append([label, count])
            
        elif 'projects' in data:
            # Comparison data
            ws_comparison = wb.create_sheet("Comparison")
            
            # Create comparison table
            headers = ["Project", "Health Score", "Total Commits", "Open Issues", "Merge Rate"]
            ws_comparison.append(headers)
            
            for pid, project_data in data['projects'].items():
                row = [
                    project_data['name'],
                    project_data.get('health_score', 0),
                    project_data.get('metrics', {}).get('commits', {}).get('total_commits', 0),
                    project_data.get('metrics', {}).get('issues', {}).get('open_issues', 0),
                    project_data.get('metrics', {}).get('merge_requests', {}).get('merge_rate', 0)
                ]
                ws_comparison.append(row)
            
            # Style headers
            for cell in ws_comparison[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report saved to: {output_path}")
        
    except ImportError:
        logger.error("pandas and openpyxl are required for Excel export")
        logger.error("Install with: pip install pandas openpyxl")
        raise


def export_to_html(data: dict, output_path: str):
    """Export analytics data to HTML format.
    
    Args:
        data: Analytics data
        output_path: Output file path
    """
    html_template = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>GitLab Analytics Report</title>
        <style>
            body {{
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                line-height: 1.6;
                color: #333;
                max-width: 1200px;
                margin: 0 auto;
                padding: 20px;
                background-color: #f8f9fa;
            }}
            .header {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 30px;
                border-radius: 10px;
                margin-bottom: 30px;
                text-align: center;
            }}
            .header h1 {{
                margin: 0;
                font-size: 2.5em;
            }}
            .header .subtitle {{
                margin: 10px 0 0 0;
                opacity: 0.9;
            }}
            .card {{
                background: white;
                border-radius: 10px;
                padding: 25px;
                margin-bottom: 25px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            }}
            .card h2 {{
                color: #2c3e50;
                border-bottom: 3px solid #3498db;
                padding-bottom: 10px;
                margin-top: 0;
            }}
            .metrics-grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 20px;
                margin: 20px 0;
            }}
            .metric {{
                background: #f8f9fa;
                padding: 20px;
                border-radius: 8px;
                text-align: center;
                border-left: 4px solid #3498db;
            }}
            .metric-value {{
                font-size: 2em;
                font-weight: bold;
                color: #2c3e50;
                margin-bottom: 5px;
            }}
            .metric-label {{
                color: #7f8c8d;
                font-size: 0.9em;
            }}
            .table {{
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
            }}
            .table th, .table td {{
                padding: 12px;
                text-align: left;
                border-bottom: 1px solid #ddd;
            }}
            .table th {{
                background-color: #3498db;
                color: white;
                font-weight: bold;
            }}
            .table tr:hover {{
                background-color: #f5f5f5;
            }}
            .status-active {{ color: #27ae60; font-weight: bold; }}
            .status-inactive {{ color: #e74c3c; font-weight: bold; }}
            .progress-bar {{
                background-color: #ecf0f1;
                border-radius: 10px;
                overflow: hidden;
                height: 20px;
                margin: 5px 0;
            }}
            .progress-fill {{
                height: 100%;
                background: linear-gradient(90deg, #3498db, #2ecc71);
                transition: width 0.3s ease;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>📊 GitLab Analytics Report</h1>
            <div class="subtitle">Generated on {timestamp}</div>
        </div>
        
        {content}
    </body>
    </html>
    """
    
    content = ""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    if 'project' in data:
        # Single project report
        project = data['project']
        
        # Project info card
        content += f"""
        <div class="card">
            <h2>🏷️ Project Information</h2>
            <div class="metrics-grid">
                <div class="metric">
                    <div class="metric-value">{project['name']}</div>
                    <div class="metric-label">Project Name</div>
                </div>
                <div class="metric">
                    <div class="metric-value">{project['visibility'].title()}</div>
                    <div class="metric-label">Visibility</div>
                </div>
                <div class="metric">
                    <div class="metric-value">{project.get('default_branch', 'N/A')}</div>
                    <div class="metric-label">Default Branch</div>
                </div>
            </div>
            <p><strong>Path:</strong> {project['path']}</p>
            <p><strong>Created:</strong> {project['created_at']}</p>
            <p><strong>Last Activity:</strong> {project['last_activity_at']}</p>
        </div>
        """
        
        # Metrics cards
        if 'commits' in data:
            commits = data['commits']
            content += f"""
            <div class="card">
                <h2>📝 Commit Statistics</h2>
                <div class="metrics-grid">
                    <div class="metric">
                        <div class="metric-value">{commits.get('total', 0)}</div>
                        <div class="metric-label">Total Commits (30 days)</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value">{commits.get('average_per_day', 0):.1f}</div>
                        <div class="metric-label">Average per Day</div>
                    </div>
                </div>
            """
            
            if commits.get('by_author'):
                content += """
                <h3>Top Contributors</h3>
                <table class="table">
                    <thead>
                        <tr><th>Author</th><th>Commits</th><th>Percentage</th></tr>
                    </thead>
                    <tbody>
                """
                total_commits = sum(commits['by_author'].values())
                for author, count in sorted(commits['by_author'].items(), key=lambda x: x[1], reverse=True)[:10]:
                    percentage = (count / total_commits * 100) if total_commits > 0 else 0
                    content += f"""
                        <tr>
                            <td>{author}</td>
                            <td>{count}</td>
                            <td>
                                <div class="progress-bar">
                                    <div class="progress-fill" style="width: {percentage}%"></div>
                                </div>
                                {percentage:.1f}%
                            </td>
                        </tr>
                    """
                content += "</tbody></table>"
            content += "</div>"
        
        if 'issues' in data:
            issues = data['issues']
            closure_rate = issues.get('closure_rate', 0) * 100
            content += f"""
            <div class="card">
                <h2>🐛 Issue Statistics</h2>
                <div class="metrics-grid">
                    <div class="metric">
                        <div class="metric-value">{issues.get('total', 0)}</div>
                        <div class="metric-label">Total Issues</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value">{issues.get('open', 0)}</div>
                        <div class="metric-label">Open Issues</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value">{issues.get('closed', 0)}</div>
                        <div class="metric-label">Closed Issues</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value">{closure_rate:.1f}%</div>
                        <div class="metric-label">Closure Rate</div>
                    </div>
                </div>
            </div>
            """
        
        if 'branches' in data:
            branches = data['branches']
            content += f"""
            <div class="card">
                <h2>🌿 Branch Statistics</h2>
                <div class="metrics-grid">
                    <div class="metric">
                        <div class="metric-value">{branches.get('total', 0)}</div>
                        <div class="metric-label">Total Branches</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value">{branches.get('active', 0)}</div>
                        <div class="metric-label">Active Branches</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value">{branches.get('stale', 0)}</div>
                        <div class="metric-label">Stale Branches</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value">{branches.get('protected', 0)}</div>
                        <div class="metric-label">Protected Branches</div>
                    </div>
                </div>
            </div>
            """
    
    elif 'projects' in data:
        # Comparison report
        content += """
        <div class="card">
            <h2>📊 Project Comparison</h2>
            <table class="table">
                <thead>
                    <tr>
                        <th>Project</th>
                        <th>Health Score</th>
                        <th>Commits</th>
                        <th>Issues</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for pid, project_data in data['projects'].items():
            health_score = project_data.get('health_score', 0)
            status_class = "status-active" if health_score > 70 else "status-inactive"
            content += f"""
                <tr>
                    <td><strong>{project_data.get('name', pid)}</strong></td>
                    <td>{health_score}/100</td>
                    <td>{project_data.get('metrics', {}).get('commits', {}).get('total_commits', 0)}</td>
                    <td>{project_data.get('metrics', {}).get('issues', {}).get('open_issues', 0)}</td>
                    <td class="{status_class}">{'Healthy' if health_score > 70 else 'Needs Attention'}</td>
                </tr>
            """
        
        content += "</tbody></table></div>"
    
    # Generate final HTML
    html_content = html_template.format(
        timestamp=timestamp,
        content=content
    )
    
    # Save to file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    logger.info(f"HTML report saved to: {output_path}")


def main():
    """Main function."""
    parser = argparse.ArgumentParser(
        description="Export GitLab analytics to various formats"
    )
    
    # Target selection
    parser.add_argument(
        'target',
        help='Project ID, group name, or "compare:id1,id2,id3" for comparison'
    )
    
    # Output options
    parser.add_argument(
        '--output', '-o',
        required=True,
        help='Output file path'
    )
    parser.add_argument(
        '--format', '-f',
        choices=['excel', 'csv', 'json', 'html'],
        default='excel',
        help='Output format (default: excel)'
    )
    
    # Analytics options
    parser.add_argument(
        '--trends',
        action='store_true',
        help='Include trend analysis'
    )
    parser.add_argument(
        '--days',
        type=int,
        default=30,
        help='Number of days for analysis (default: 30)'
    )
    
    # Configuration
    parser.add_argument(
        '--config', '-c',
        help='Path to configuration file'
    )
    parser.add_argument(
        '--log-level',
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        default='INFO',
        help='Logging level'
    )
    
    args = parser.parse_args()
    
    # Load configuration
    config = Config(args.config)
    
    # Setup logging
    setup_logging(
        config.get_log_config(),
        console_level=args.log_level
    )
    
    try:
        # Validate configuration
        config.validate()
        
        # Create GitLab client
        gitlab_config = config.get_gitlab_config()
        client = GitLabClient(
            url=gitlab_config['url'],
            token=gitlab_config['token'],
            config=gitlab_config
        )
        
        # Create analytics services
        analytics = GitLabAnalytics(client)
        advanced = AdvancedAnalytics(client)
        
        print(f"{Colors.BOLD}Generating analytics export...{Colors.RESET}")
        
        # Determine target type
        if args.target.startswith('compare:'):
            # Comparison mode
            project_ids = args.target.split(':', 1)[1].split(',')
            logger.info(f"Comparing projects: {project_ids}")
            data = advanced.compare_projects(project_ids)
        else:
            # Single project or group
            if args.target.isdigit():
                # Project ID
                logger.info(f"Analyzing project: {args.target}")
                if args.trends:
                    data = advanced.get_project_trends(args.target, days=args.days)
                else:
                    data = analytics.get_project_metrics(args.target)
            else:
                # Try as group name
                group = client.search_group_by_name(args.target)
                if group:
                    logger.info(f"Analyzing group: {args.target}")
                    data = analytics.get_group_metrics(group['id'])
                else:
                    # Try as project path
                    logger.info(f"Analyzing project: {args.target}")
                    data = analytics.get_project_metrics(args.target)
        
        # Export based on format
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        if args.format == 'excel':
            if not output_path.suffix:
                output_path = output_path.with_suffix('.xlsx')
            export_to_excel(data, str(output_path))
        
        elif args.format == 'csv':
            # Simple CSV export of main metrics
            import csv
            
            if not output_path.suffix:
                output_path = output_path.with_suffix('.csv')
            
            with open(output_path, 'w', newline='') as f:
                writer = csv.writer(f)
                
                if 'project' in data:
                    # Single project
                    writer.writerow(['Metric', 'Value'])
                    writer.writerow(['Project Name', data['project']['name']])
                    writer.writerow(['Total Commits', data.get('commits', {}).get('total', 0)])
                    writer.writerow(['Open Issues', data.get('issues', {}).get('open', 0)])
                    writer.writerow(['Total Branches', data.get('branches', {}).get('total', 0)])
                    writer.writerow(['Contributors', data.get('contributors', {}).get('total', 0)])
                else:
                    # Group or comparison - write as table
                    writer.writerow(['Project', 'Commits', 'Issues', 'Branches'])
                    # Add data rows based on structure
        
        elif args.format == 'json':
            if not output_path.suffix:
                output_path = output_path.with_suffix('.json')
            
            with open(output_path, 'w') as f:
                json.dump(data, f, indent=2, default=str)
        
        elif args.format == 'html':
            if not output_path.suffix:
                output_path = output_path.with_suffix('.html')
            export_to_html(data, str(output_path))
        
        print(f"{Colors.GREEN}Export saved to: {output_path}{Colors.RESET}")
        
        return 0
        
    except Exception as e:
        logger.error(f"Failed to export analytics: {e}", exc_info=True)
        print(f"{Colors.RED}Error: {e}{Colors.RESET}")
        return 1


if __name__ == "__main__":
    sys.exit(main())